import sys
import socket
import os
import urllib
import whois
import subprocess 
import nmap 
import json 
import urllib
import openpyxl
import requests
wb_output = openpyxl.Workbook()
workSheet_output = wb_output.active
workSheet_output.title = "dnsResolve_Output"

workSheet_output.cell(row=1,column=1).value = "Domain Name"
workSheet_output.cell(row=1,column=2).value = "IP"
workSheet_output.cell(row=1,column=3).value = "Port Scan"
workSheet_output.cell(row=1,column=4).value = "HTTP Response Code"
workSheet_output.cell(row=1,column=5).value = "LOCATION"
workSheet_output.cell(row=1,column=6).value= "CNAME"
workSheet_output.cell(row=1,column=7).value = "NSLookUp"

i = 2
f = open("list.txt","r")
for u in f.read().split('\n'):
	
	ip_addr = socket.gethostbyname(u)
	workSheet_output.cell(row=i,column=1).value = u
	print "=============================================================="
	print "DNS :"+u+" iP Address: "+ip_addr
	workSheet_output.cell(row=i,column=2).value = ip_addr

	process = subprocess.Popen(["nslookup",u],stdout=subprocess.PIPE,shell=True)
	nslookup_output = process.communicate()[0]
	workSheet_output.cell(row=i,column=7).value = nslookup_output
	print "NSLookup : "+str(nslookup_output)
	process1 = subprocess.Popen(["nmap",u],shell=True,stdout=subprocess.PIPE)
	nmap_output = process1.communicate()[0]
	print "Nmap Scan: "+str(nmap_output)
	workSheet_output.cell(row=i,column=3).value = nmap_output
	process2 = subprocess.Popen(["curl","ipinfo.io/"+str(ip_addr)+"/city"],stdout=subprocess.PIPE,shell=True)
	Location_Output = process2.communicate()[0]
	print "Location : "+Location_Output
	process3 = subprocess.Popen(["curl","ipinfo.io/"+str(ip_addr)+"/json"],stdout=subprocess.PIPE,shell=True)
	IPInfo_Output = process3.communicate()[0]
	print "IP Json Data: "+IPInfo_Output
	json_data=json.loads(IPInfo_Output)
	print "CNAME: "+str(json_data['hostname'])
	workSheet_output.cell(row=i,column=6).value = str(json_data['hostname'])
	workSheet_output.cell(row=i,column=5).value = Location_Output
	httpURL = str("https://")+u
	
	i=i+1

	wb_output.save("dnsResolve.xlsx")
