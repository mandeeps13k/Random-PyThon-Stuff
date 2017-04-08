# A Simple Script to dump data from JustDial using two Keywords and Dump the output in an Excel File

import urllib2
import sys
from bs4 import BeautifulSoup
import json
import openpyxl
wb_output = openpyxl.Workbook()
worksheet_output = wb_output.active
worksheet_output.title = "Just Dial Data Dump"
worksheet_output.cell(row=1,column=1).value = "Buisness Name"
worksheet_output.cell(row=1,column=2).value = "Street Address"
worksheet_output.cell(row=1,column=3).value = "Area"
worksheet_output.cell(row=1,column=4).value = "Pincode"
worksheet_output.cell(row=1,column=5).value = "Phone Numbers"

worksheet_output.column_dimensions['A'].width = 50
worksheet_output.column_dimensions['B'].width = 50
worksheet_output.column_dimensions['C'].width = 50
worksheet_output.column_dimensions['D'].width = 50
worksheet_output.column_dimensions['E'].width = 50
keyword1 = sys.argv[1]
keyword2 = sys.argv[2]
httpURL = "http://justdial.com/"+str(keyword1)+"/"+str(keyword2)
print "Loading Url : "+httpURL
headers = {'User-Agent':'Mozilla/5.0'}
req = urllib2.Request(httpURL,None,headers)
html = urllib2.urlopen(req).read()
bs = BeautifulSoup(html,"lxml")
for i in range(0,10):
	allLinks = bs.find_all('a',id='newphoto'+str(i))
	print allLinks[0]['href']
	LinkURL = allLinks[0]['href']
	headers = {'User-Agent':'Mozilla/5.0'}
	req = urllib2.Request(LinkURL,None,headers)
	html2 = urllib2.urlopen(req).read()
	bs1 = BeautifulSoup(html2,"lxml")
	bs1.find_all('script')[1]
	print bs1.find_all('script')[1]
	script = bs1.find_all('script')[1].string
	data = json.loads(script)
	print data
	print "Buisness Name  : "+str(data["name"])
	worksheet_output.cell(row=i+2,column=1).value = str(data["name"])
	print "Telephone: "+str(data["telephone"])
	worksheet_output.cell(row=i+2,column=5).value = str(data["telephone"])
	print "StreetAddress: "+str(data["address"]["streetAddress"])
	worksheet_output.cell(row=i+2,column=2).value = str(data["address"]["streetAddress"])
	print "Postal Code: "+str(data["address"]["postalCode"])
	worksheet_output.cell(row=i+2,column=4).value = str(data["address"]["postalCode"])
	print "Address Locality: "+str(data["address"]["addressLocality"])
	worksheet_output.cell(row=i+2,column=3).value = str(data["address"]["addressLocality"])

wb_output.save("JustDialDump.xlsx")
